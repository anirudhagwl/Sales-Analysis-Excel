import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel

df = pd.read_csv('data/Sales_Orders.csv')
df['Order Date'] = pd.to_datetime(df['Order Date'], format='%d/%m/%Y')
df['Ship Date'] = pd.to_datetime(df['Ship Date'], format='%d/%m/%Y')
df['Year'] = df['Order Date'].dt.year
df['Month'] = df['Order Date'].dt.month
df['Delivery_Days'] = (df['Ship Date'] - df['Order Date']).dt.days
N = len(df)
LAST = N + 1

wb = Workbook()

DARK_BLUE = PatternFill('solid', fgColor='1F4E79')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')
KPI_FILL = PatternFill('solid', fgColor='2E75B6')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Arial', bold=True, color='1F4E79', size=14)
BOLD_FONT = Font(name='Arial', bold=True, size=10)
DATA_FONT = Font(name='Arial', size=10)
RED_FONT = Font(name='Arial', size=10, color='FF0000', bold=True)
CURRENCY_FMT = '$#,##0.00'
NUM_FMT = '#,##0'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Chart heights in rows (openpyxl cm -> ~rows). Use these to compute safe gaps.
CHART_ROWS_TALL = 28   # for height=14
CHART_ROWS_MED = 25    # for height=12-13
CHART_ROWS_SHORT = 22  # for height=11

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = DARK_BLUE
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_data_rows(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = WHITE_FILL if (r - start_row) % 2 == 0 else LIGHT_GRAY

def add_title(ws, title, row=1, col=1):
    ws.cell(row=row, column=col, value=title).font = TITLE_FONT

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def fmt_cells(ws, rows, col, fmt):
    for r in rows:
        ws.cell(row=r, column=col).number_format = fmt

# =====================================================================
# SHEET 1: RAW DATA
# =====================================================================
ws = wb.active
ws.title = 'Raw Data'
raw_cols = ['Row ID','Order ID','Order Date','Ship Date','Ship Mode','Customer ID',
            'Customer Name','Segment','Country','City','State','Postal Code','Region',
            'Category','Sub-Category','Product Name','Sales','Quantity','Discount',
            'Profit','Delivery Days','Profit Margin']
for c, h in enumerate(raw_cols, 1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, len(raw_cols))

for i, (_, row) in enumerate(df.iterrows(), 2):
    ws.cell(row=i, column=1, value=row['Row ID'])
    ws.cell(row=i, column=2, value=row['Order ID'])
    ws.cell(row=i, column=3, value=row['Order Date'].strftime('%Y-%m-%d'))
    ws.cell(row=i, column=4, value=row['Ship Date'].strftime('%Y-%m-%d'))
    ws.cell(row=i, column=5, value=row['Ship Mode'])
    ws.cell(row=i, column=6, value=row['Customer ID'])
    ws.cell(row=i, column=7, value=row['Customer Name'])
    ws.cell(row=i, column=8, value=row['Segment'])
    ws.cell(row=i, column=9, value=row['Country'])
    ws.cell(row=i, column=10, value=row['City'])
    ws.cell(row=i, column=11, value=row['State'])
    ws.cell(row=i, column=12, value=row['Postal Code'])
    ws.cell(row=i, column=13, value=row['Region'])
    ws.cell(row=i, column=14, value=row['Category'])
    ws.cell(row=i, column=15, value=row['Sub-Category'])
    ws.cell(row=i, column=16, value=row['Product Name'])
    ws.cell(row=i, column=17, value=row['Sales'])
    ws.cell(row=i, column=18, value=row['Quantity'])
    ws.cell(row=i, column=19, value=row['Discount'])
    ws.cell(row=i, column=20, value=row['Profit'])
    ws.cell(row=i, column=21, value=f'=D{i}-C{i}')
    ws.cell(row=i, column=22, value=f'=IF(Q{i}=0,0,T{i}/Q{i})')
    ws.cell(row=i, column=17).number_format = CURRENCY_FMT
    ws.cell(row=i, column=20).number_format = CURRENCY_FMT
    ws.cell(row=i, column=19).number_format = PCT_FMT
    ws.cell(row=i, column=22).number_format = PCT_FMT

ws.freeze_panes = 'A2'
set_col_widths(ws, [8,16,12,12,14,14,22,12,14,16,16,10,10,14,16,40,12,10,10,12,12,12])
print(f'Raw Data: {N} rows, formulas in cols U (Delivery Days) and V (Profit Margin)')

# =====================================================================
# SHEET 2: SUMMARY STATISTICS
# =====================================================================
ws2 = wb.create_sheet('Summary Statistics')
add_title(ws2, 'Descriptive Statistics — All Values Computed via Excel Formulas', 1, 1)

ref_map = {'Sales ($)': 'Q', 'Profit ($)': 'T', 'Quantity': 'R', 'Discount': 'S', 'Delivery Days': 'U'}
stat_headers = ['Statistic'] + list(ref_map.keys())
for c, h in enumerate(stat_headers, 1):
    ws2.cell(row=3, column=c, value=h)
style_header(ws2, 3, len(stat_headers))

stats_list = [
    ('Count', 'COUNT'), ('Sum', 'SUM'), ('Mean', 'AVERAGE'), ('Median', 'MEDIAN'),
    ('Std Deviation', 'STDEV'), ('Min', 'MIN'), ('Max', 'MAX'),
    ('Q1 (25th Pctl)', 'PERCENTILE_25'), ('Q3 (75th Pctl)', 'PERCENTILE_75'),
    ('Variance', 'VAR'),
]
for r, (label, func) in enumerate(stats_list, 4):
    ws2.cell(row=r, column=1, value=label)
    for c, (_, col_letter) in enumerate(ref_map.items(), 2):
        rng = f"'Raw Data'!{col_letter}2:{col_letter}{LAST}"
        if func == 'PERCENTILE_25':
            ws2.cell(row=r, column=c, value=f'=PERCENTILE({rng},0.25)')
        elif func == 'PERCENTILE_75':
            ws2.cell(row=r, column=c, value=f'=PERCENTILE({rng},0.75)')
        else:
            ws2.cell(row=r, column=c, value=f'={func}({rng})')

r_iqr = 14
ws2.cell(row=r_iqr, column=1, value='IQR (Q3 - Q1)')
for c in range(2, len(stat_headers) + 1):
    cl = get_column_letter(c)
    ws2.cell(row=r_iqr, column=c, value=f'={cl}11-{cl}10')

style_data_rows(ws2, 4, r_iqr, len(stat_headers))
for r in range(4, r_iqr + 1):
    ws2.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws2.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws2.cell(row=r, column=5).number_format = PCT_FMT

# Segment breakdown
seg_start = r_iqr + 3
add_title(ws2, 'Performance by Customer Segment', seg_start, 1)
seg_h = ['Segment', 'Total Sales', 'Total Profit', 'Profit Margin', 'Order Count']
for c, h in enumerate(seg_h, 1):
    ws2.cell(row=seg_start + 2, column=c, value=h)
style_header(ws2, seg_start + 2, len(seg_h))

segments = ['Consumer', 'Corporate', 'Home Office']
for r, seg in enumerate(segments, seg_start + 3):
    ws2.cell(row=r, column=1, value=seg)
    ws2.cell(row=r, column=2, value=f"=SUMIF('Raw Data'!H2:H{LAST},A{r},'Raw Data'!Q2:Q{LAST})")
    ws2.cell(row=r, column=3, value=f"=SUMIF('Raw Data'!H2:H{LAST},A{r},'Raw Data'!T2:T{LAST})")
    ws2.cell(row=r, column=4, value=f'=IF(B{r}=0,0,C{r}/B{r})')
    ws2.cell(row=r, column=5, value=f"=COUNTIF('Raw Data'!H2:H{LAST},A{r})")
    ws2.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws2.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws2.cell(row=r, column=4).number_format = PCT_FMT

seg_data_end = seg_start + 5
tr = seg_data_end + 1
ws2.cell(row=tr, column=1, value='TOTAL').font = BOLD_FONT
for c in [2, 3, 5]:
    cl = get_column_letter(c)
    ws2.cell(row=tr, column=c, value=f'=SUM({cl}{seg_start+3}:{cl}{seg_data_end})')
    ws2.cell(row=tr, column=c).font = BOLD_FONT
ws2.cell(row=tr, column=4, value=f'=IF(B{tr}=0,0,C{tr}/B{tr})')
ws2.cell(row=tr, column=2).number_format = CURRENCY_FMT
ws2.cell(row=tr, column=3).number_format = CURRENCY_FMT
ws2.cell(row=tr, column=4).number_format = PCT_FMT
style_data_rows(ws2, seg_start + 3, tr, len(seg_h))
set_col_widths(ws2, [18, 16, 16, 14, 14, 14])

# =====================================================================
# SHEET 3: REVENUE ANALYSIS
# =====================================================================
ws3 = wb.create_sheet('Revenue Analysis')
add_title(ws3, 'Revenue Trend Analysis', 1, 1)

years = sorted(df['Year'].unique())
rev_h = ['Year', 'Total Revenue', 'Total Profit', 'Total Orders', 'Avg Order Value', 'Profit Margin', 'YoY Growth']
for c, h in enumerate(rev_h, 1):
    ws3.cell(row=3, column=c, value=h)
style_header(ws3, 3, len(rev_h))

for r, year in enumerate(years, 4):
    ws3.cell(row=r, column=1, value=year)
    ws3.cell(row=r, column=2, value=f"=SUMPRODUCT((LEFT('Raw Data'!C2:C{LAST},4)=TEXT(A{r},\"0000\"))*'Raw Data'!Q2:Q{LAST})")
    ws3.cell(row=r, column=3, value=f"=SUMPRODUCT((LEFT('Raw Data'!C2:C{LAST},4)=TEXT(A{r},\"0000\"))*'Raw Data'!T2:T{LAST})")
    ws3.cell(row=r, column=4, value=f"=SUMPRODUCT((LEFT('Raw Data'!C2:C{LAST},4)=TEXT(A{r},\"0000\"))*1)")
    ws3.cell(row=r, column=5, value=f'=IF(D{r}=0,0,B{r}/D{r})')
    ws3.cell(row=r, column=6, value=f'=IF(B{r}=0,0,C{r}/B{r})')
    if r == 4:
        ws3.cell(row=r, column=7, value='N/A')
    else:
        ws3.cell(row=r, column=7, value=f'=IF(B{r-1}=0,0,(B{r}-B{r-1})/B{r-1})')
        ws3.cell(row=r, column=7).number_format = PCT_FMT
    for c in [2, 3, 5]:
        ws3.cell(row=r, column=c).number_format = CURRENCY_FMT
    ws3.cell(row=r, column=6).number_format = PCT_FMT

yr_end = 3 + len(years)
tr = yr_end + 1
ws3.cell(row=tr, column=1, value='TOTAL').font = BOLD_FONT
for c in [2, 3, 4]:
    cl = get_column_letter(c)
    ws3.cell(row=tr, column=c, value=f'=SUM({cl}4:{cl}{yr_end})')
    ws3.cell(row=tr, column=c).font = BOLD_FONT
ws3.cell(row=tr, column=5, value=f'=IF(D{tr}=0,0,B{tr}/D{tr})')
ws3.cell(row=tr, column=6, value=f'=IF(B{tr}=0,0,C{tr}/B{tr})')
ws3.cell(row=tr, column=2).number_format = CURRENCY_FMT
ws3.cell(row=tr, column=3).number_format = CURRENCY_FMT
ws3.cell(row=tr, column=5).number_format = CURRENCY_FMT
ws3.cell(row=tr, column=6).number_format = PCT_FMT
style_data_rows(ws3, 4, tr, len(rev_h))
set_col_widths(ws3, [10, 16, 16, 14, 16, 14, 14])

# Bar chart: place at row tr+3, takes ~25 rows
chart_start_1 = tr + 3
chart = BarChart()
chart.type = "col"; chart.title = "Yearly Revenue & Profit"; chart.style = 10
chart.width = 20; chart.height = 13; chart.y_axis.title = "Amount ($)"
chart.add_data(Reference(ws3, min_col=2, min_row=3, max_row=yr_end), titles_from_data=True)
chart.add_data(Reference(ws3, min_col=3, min_row=3, max_row=yr_end), titles_from_data=True)
chart.set_categories(Reference(ws3, min_col=1, min_row=4, max_row=yr_end))
ws3.add_chart(chart, f"A{chart_start_1}")

# Monthly breakdown: starts AFTER chart (chart_start + 27 rows buffer)
mon_start = chart_start_1 + CHART_ROWS_MED + 3
add_title(ws3, 'Monthly Revenue Breakdown', mon_start, 1)
months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
mon_h = ['Year'] + months + ['Year Total']
for c, h in enumerate(mon_h, 1):
    ws3.cell(row=mon_start + 2, column=c, value=h)
style_header(ws3, mon_start + 2, len(mon_h))

monthly = df.groupby([df['Year'], df['Month']])['Sales'].sum()
for yr_idx, year in enumerate(years):
    r = mon_start + 3 + yr_idx
    ws3.cell(row=r, column=1, value=int(year))
    for m in range(1, 13):
        val = monthly.get((year, m), 0)
        ws3.cell(row=r, column=m + 1, value=round(val, 2))
        ws3.cell(row=r, column=m + 1).number_format = CURRENCY_FMT
    ws3.cell(row=r, column=14, value=f'=SUM(B{r}:M{r})')
    ws3.cell(row=r, column=14).number_format = CURRENCY_FMT
    ws3.cell(row=r, column=14).font = BOLD_FONT

mon_end = mon_start + 3 + len(years) - 1
style_data_rows(ws3, mon_start + 3, mon_end, len(mon_h))

# Line chart: after monthly table
line_start = mon_end + 3
line_chart = LineChart()
line_chart.title = "Monthly Revenue by Year"; line_chart.style = 10
line_chart.width = 24; line_chart.height = 14; line_chart.y_axis.title = "Revenue ($)"
cats = Reference(ws3, min_col=2, min_row=mon_start + 2, max_col=13)
for yr_idx, year in enumerate(years):
    data = Reference(ws3, min_col=2, max_col=13, min_row=mon_start + 3 + yr_idx)
    line_chart.add_data(data, from_rows=True, titles_from_data=False)
    line_chart.series[yr_idx].tx = SeriesLabel(v=str(int(year)))
line_chart.set_categories(cats)
ws3.add_chart(line_chart, f"A{line_start}")

# =====================================================================
# SHEET 4: PROFITABILITY ANALYSIS
# =====================================================================
ws4 = wb.create_sheet('Profitability Analysis')
add_title(ws4, 'Profitability Analysis', 1, 1)

# Category table
add_title(ws4, 'By Category', 3, 1)
cat_h = ['Category', 'Total Sales', 'Total Profit', 'Profit Margin', 'Units Sold']
for c, h in enumerate(cat_h, 1):
    ws4.cell(row=5, column=c, value=h)
style_header(ws4, 5, len(cat_h))

categories = ['Technology', 'Furniture', 'Office Supplies']
for r, cat in enumerate(categories, 6):
    ws4.cell(row=r, column=1, value=cat)
    ws4.cell(row=r, column=2, value=f"=SUMIF('Raw Data'!N2:N{LAST},A{r},'Raw Data'!Q2:Q{LAST})")
    ws4.cell(row=r, column=3, value=f"=SUMIF('Raw Data'!N2:N{LAST},A{r},'Raw Data'!T2:T{LAST})")
    ws4.cell(row=r, column=4, value=f'=IF(B{r}=0,0,C{r}/B{r})')
    ws4.cell(row=r, column=5, value=f"=SUMIF('Raw Data'!N2:N{LAST},A{r},'Raw Data'!R2:R{LAST})")
    ws4.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws4.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws4.cell(row=r, column=4).number_format = PCT_FMT

tr = 9
ws4.cell(row=tr, column=1, value='TOTAL').font = BOLD_FONT
for c in [2, 3, 5]:
    cl = get_column_letter(c)
    ws4.cell(row=tr, column=c, value=f'=SUM({cl}6:{cl}8)')
    ws4.cell(row=tr, column=c).font = BOLD_FONT
ws4.cell(row=tr, column=4, value=f'=IF(B{tr}=0,0,C{tr}/B{tr})')
ws4.cell(row=tr, column=2).number_format = CURRENCY_FMT
ws4.cell(row=tr, column=3).number_format = CURRENCY_FMT
ws4.cell(row=tr, column=4).number_format = PCT_FMT
style_data_rows(ws4, 6, tr, len(cat_h))

# Pie chart: BELOW table, not to the right
pie_start = tr + 3
pie = PieChart()
pie.title = "Sales by Category"; pie.style = 10; pie.width = 14; pie.height = 11
pie.add_data(Reference(ws4, min_col=2, min_row=6, max_row=8))
pie.set_categories(Reference(ws4, min_col=1, min_row=6, max_row=8))
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True
ws4.add_chart(pie, f"A{pie_start}")

# Sub-category: after pie chart
sub_start = pie_start + CHART_ROWS_SHORT + 3
add_title(ws4, 'By Sub-Category', sub_start, 1)
subcat_h = ['Sub-Category', 'Total Sales', 'Total Profit', 'Profit Margin']
for c, h in enumerate(subcat_h, 1):
    ws4.cell(row=sub_start + 2, column=c, value=h)
style_header(ws4, sub_start + 2, len(subcat_h))

subcats = sorted(df['Sub-Category'].unique())
for r, sc in enumerate(subcats, sub_start + 3):
    ws4.cell(row=r, column=1, value=sc)
    ws4.cell(row=r, column=2, value=f"=SUMIF('Raw Data'!O2:O{LAST},A{r},'Raw Data'!Q2:Q{LAST})")
    ws4.cell(row=r, column=3, value=f"=SUMIF('Raw Data'!O2:O{LAST},A{r},'Raw Data'!T2:T{LAST})")
    ws4.cell(row=r, column=4, value=f'=IF(B{r}=0,0,C{r}/B{r})')
    ws4.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws4.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws4.cell(row=r, column=4).number_format = PCT_FMT

sub_end = sub_start + 2 + len(subcats)
style_data_rows(ws4, sub_start + 3, sub_end, len(subcat_h))

# Sub-category bar chart: after sub-category table
bar_start = sub_end + 3
bar = BarChart()
bar.type = "col"; bar.title = "Profit by Sub-Category"; bar.style = 10
bar.width = 22; bar.height = 14
bar.add_data(Reference(ws4, min_col=3, min_row=sub_start + 2, max_row=sub_end), titles_from_data=True)
bar.set_categories(Reference(ws4, min_col=1, min_row=sub_start + 3, max_row=sub_end))
ws4.add_chart(bar, f"A{bar_start}")

set_col_widths(ws4, [18, 16, 16, 14, 14])

# =====================================================================
# SHEET 5: CUSTOMER SEGMENTATION — fully formula-based RFM
# =====================================================================
ws5 = wb.create_sheet('Customer Segmentation')
add_title(ws5, 'RFM Customer Segmentation Analysis', 1, 1)
ws5.cell(row=2, column=1, value='All scores computed via Excel formulas (PERCENTRANK, IF). Click any cell to verify.').font = Font(name='Arial', size=9, italic=True, color='666666')

# Compute per-customer raw data from pandas (aggregates needed)
max_date = df['Order Date'].max()
rfm = df.groupby('Customer ID').agg({
    'Order Date': lambda x: (max_date - x.max()).days,
    'Order ID': 'nunique',
    'Sales': 'sum'
}).reset_index()
rfm.columns = ['Customer ID', 'Recency', 'Frequency', 'Monetary']
cust_names = df.drop_duplicates('Customer ID').set_index('Customer ID')['Customer Name'].to_dict()
rfm['Customer Name'] = rfm['Customer ID'].map(cust_names)
rfm = rfm.sort_values('Monetary', ascending=False)
n_cust = len(rfm)

rfm_h = ['Customer ID', 'Customer Name', 'Recency (days)', 'Frequency', 'Monetary ($)',
          'R Score', 'F Score', 'M Score', 'RFM Score', 'Segment']
for c, h in enumerate(rfm_h, 1):
    ws5.cell(row=4, column=c, value=h)
style_header(ws5, 4, len(rfm_h))

CUST_START = 5
CUST_END = 4 + n_cust
for r, (_, row) in enumerate(rfm.iterrows(), CUST_START):
    ws5.cell(row=r, column=1, value=row['Customer ID'])
    ws5.cell(row=r, column=2, value=row['Customer Name'])
    ws5.cell(row=r, column=3, value=int(row['Recency']))
    ws5.cell(row=r, column=4, value=int(row['Frequency']))
    ws5.cell(row=r, column=5, value=round(row['Monetary'], 2))
    ws5.cell(row=r, column=5).number_format = CURRENCY_FMT

    # R Score: Lower recency = better = higher score (4=best, 1=worst)
    # PERCENTRANK gives 0 for lowest value, 1 for highest
    # Low recency -> low percentrank -> high score
    rec_rng = f'$C${CUST_START}:$C${CUST_END}'
    ws5.cell(row=r, column=6, value=(
        f'=IF(PERCENTRANK({rec_rng},C{r})<=0.25,4,'
        f'IF(PERCENTRANK({rec_rng},C{r})<=0.5,3,'
        f'IF(PERCENTRANK({rec_rng},C{r})<=0.75,2,1)))'
    ))

    # F Score: Higher frequency = better = higher score
    freq_rng = f'$D${CUST_START}:$D${CUST_END}'
    ws5.cell(row=r, column=7, value=(
        f'=IF(PERCENTRANK({freq_rng},D{r})>=0.75,4,'
        f'IF(PERCENTRANK({freq_rng},D{r})>=0.5,3,'
        f'IF(PERCENTRANK({freq_rng},D{r})>=0.25,2,1)))'
    ))

    # M Score: Higher monetary = better = higher score
    mon_rng = f'$E${CUST_START}:$E${CUST_END}'
    ws5.cell(row=r, column=8, value=(
        f'=IF(PERCENTRANK({mon_rng},E{r})>=0.75,4,'
        f'IF(PERCENTRANK({mon_rng},E{r})>=0.5,3,'
        f'IF(PERCENTRANK({mon_rng},E{r})>=0.25,2,1)))'
    ))

    # RFM Score = R + F + M
    ws5.cell(row=r, column=9, value=f'=F{r}+G{r}+H{r}')

    # Segment based on RFM Score
    ws5.cell(row=r, column=10, value=(
        f'=IF(I{r}>=10,"Champions",'
        f'IF(I{r}>=8,"Loyal Customers",'
        f'IF(I{r}>=6,"Potential Loyalists",'
        f'IF(I{r}>=4,"At Risk","Lost"))))'
    ))

style_data_rows(ws5, CUST_START, CUST_END, len(rfm_h))

# Segment summary with COUNTIF/SUMIF formulas referencing above
seg_start = CUST_END + 3
add_title(ws5, 'Segment Summary', seg_start, 1)
seg_h2 = ['Segment', 'Customer Count', '% of Customers', 'Total Revenue', 'Avg Revenue/Customer']
for c, h in enumerate(seg_h2, 1):
    ws5.cell(row=seg_start + 2, column=c, value=h)
style_header(ws5, seg_start + 2, len(seg_h2))

seg_names = ['Champions', 'Loyal Customers', 'Potential Loyalists', 'At Risk', 'Lost']
seg_data_start = seg_start + 3
seg_data_end = seg_data_start + len(seg_names) - 1
for r, seg in enumerate(seg_names, seg_data_start):
    ws5.cell(row=r, column=1, value=seg)
    ws5.cell(row=r, column=2, value=f'=COUNTIF(J{CUST_START}:J{CUST_END},A{r})')
    total_cust_cell = f'SUM(B{seg_data_start}:B{seg_data_end})'
    ws5.cell(row=r, column=3, value=f'=IF({total_cust_cell}=0,0,B{r}/{total_cust_cell})')
    ws5.cell(row=r, column=3).number_format = PCT_FMT
    ws5.cell(row=r, column=4, value=f'=SUMIF(J{CUST_START}:J{CUST_END},A{r},E{CUST_START}:E{CUST_END})')
    ws5.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws5.cell(row=r, column=5, value=f'=IF(B{r}=0,0,D{r}/B{r})')
    ws5.cell(row=r, column=5).number_format = CURRENCY_FMT

tr = seg_data_end + 1
ws5.cell(row=tr, column=1, value='TOTAL').font = BOLD_FONT
ws5.cell(row=tr, column=2, value=f'=SUM(B{seg_data_start}:B{seg_data_end})')
ws5.cell(row=tr, column=2).font = BOLD_FONT
ws5.cell(row=tr, column=4, value=f'=SUM(D{seg_data_start}:D{seg_data_end})')
ws5.cell(row=tr, column=4).number_format = CURRENCY_FMT
ws5.cell(row=tr, column=4).font = BOLD_FONT
style_data_rows(ws5, seg_data_start, tr, len(seg_h2))

# Pie chart: below segment summary
pie_start_5 = tr + 3
pie2 = PieChart()
pie2.title = "Customer Segment Distribution"; pie2.style = 10; pie2.width = 14; pie2.height = 11
pie2.add_data(Reference(ws5, min_col=2, min_row=seg_data_start, max_row=seg_data_end))
pie2.set_categories(Reference(ws5, min_col=1, min_row=seg_data_start, max_row=seg_data_end))
pie2.dataLabels = DataLabelList()
pie2.dataLabels.showPercent = True
ws5.add_chart(pie2, f"A{pie_start_5}")

set_col_widths(ws5, [14, 22, 14, 12, 14, 10, 10, 10, 10, 20])

# =====================================================================
# SHEET 6: PRODUCT PERFORMANCE
# =====================================================================
ws6 = wb.create_sheet('Product Performance')
add_title(ws6, 'Product Performance Analysis', 1, 1)

add_title(ws6, 'Top 20 Products by Revenue', 3, 1)
prod_h = ['Rank', 'Product Name', 'Total Sales', 'Total Profit', 'Profit Margin', 'Units Sold']
for c, h in enumerate(prod_h, 1):
    ws6.cell(row=5, column=c, value=h)
style_header(ws6, 5, len(prod_h))

top_prods = df.groupby('Product Name').agg({'Sales':'sum','Profit':'sum','Quantity':'sum'}).reset_index()
top20 = top_prods.nlargest(20, 'Sales')
for r, (_, row) in enumerate(top20.iterrows(), 6):
    ws6.cell(row=r, column=1, value=r - 5)
    ws6.cell(row=r, column=2, value=row['Product Name'][:65])
    ws6.cell(row=r, column=3, value=round(row['Sales'], 2))
    ws6.cell(row=r, column=4, value=round(row['Profit'], 2))
    ws6.cell(row=r, column=5, value=f'=IF(C{r}=0,0,D{r}/C{r})')
    ws6.cell(row=r, column=6, value=int(row['Quantity']))
    ws6.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws6.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws6.cell(row=r, column=5).number_format = PCT_FMT
style_data_rows(ws6, 6, 25, len(prod_h))

add_title(ws6, 'Top 10 Loss-Making Products', 28, 1)
for c, h in enumerate(prod_h, 1):
    ws6.cell(row=30, column=c, value=h)
style_header(ws6, 30, len(prod_h))

bottom10 = top_prods.nsmallest(10, 'Profit')
for r, (_, row) in enumerate(bottom10.iterrows(), 31):
    ws6.cell(row=r, column=1, value=r - 30)
    ws6.cell(row=r, column=2, value=row['Product Name'][:65])
    ws6.cell(row=r, column=3, value=round(row['Sales'], 2))
    ws6.cell(row=r, column=4, value=round(row['Profit'], 2))
    ws6.cell(row=r, column=5, value=f'=IF(C{r}=0,0,D{r}/C{r})')
    ws6.cell(row=r, column=6, value=int(row['Quantity']))
    ws6.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws6.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws6.cell(row=r, column=4).font = RED_FONT
    ws6.cell(row=r, column=5).number_format = PCT_FMT
style_data_rows(ws6, 31, 40, len(prod_h))

add_title(ws6, 'ABC Analysis (Pareto 80/20 Rule)', 43, 1)
abc = top_prods.sort_values('Sales', ascending=False).reset_index(drop=True)
abc['Cum'] = abc['Sales'].cumsum() / abc['Sales'].sum() * 100
abc['Class'] = abc['Cum'].apply(lambda x: 'A' if x <= 80 else ('B' if x <= 95 else 'C'))
abc_summary = abc.groupby('Class').agg({'Product Name':'count','Sales':'sum'}).reset_index()
abc_summary.columns = ['Class', 'Product Count', 'Total Sales']

abc_h = ['Class', 'Product Count', '% of Products', 'Total Sales', '% of Revenue']
for c, h in enumerate(abc_h, 1):
    ws6.cell(row=45, column=c, value=h)
style_header(ws6, 45, len(abc_h))

for r, (_, row) in enumerate(abc_summary.iterrows(), 46):
    ws6.cell(row=r, column=1, value=row['Class'])
    ws6.cell(row=r, column=2, value=int(row['Product Count']))
    ws6.cell(row=r, column=3, value=round(row['Product Count']/len(top_prods)*100, 1))
    ws6.cell(row=r, column=4, value=round(row['Total Sales'], 2))
    ws6.cell(row=r, column=4).number_format = CURRENCY_FMT

abc_end = 45 + len(abc_summary)
total_abc = f'SUM(D46:D{abc_end})'
for r in range(46, abc_end + 1):
    ws6.cell(row=r, column=5, value=f'=IF({total_abc}=0,0,D{r}/{total_abc})')
    ws6.cell(row=r, column=5).number_format = PCT_FMT
style_data_rows(ws6, 46, abc_end, len(abc_h))
set_col_widths(ws6, [6, 55, 14, 14, 14, 12])

# =====================================================================
# SHEET 7: REGIONAL ANALYSIS
# =====================================================================
ws7 = wb.create_sheet('Regional Analysis')
add_title(ws7, 'Regional Sales & Profit Analysis', 1, 1)

regions = ['West', 'East', 'Central', 'South']
reg_h = ['Region', 'Total Sales', 'Total Profit', 'Profit Margin', 'Order Count', 'Avg Order Value']
for c, h in enumerate(reg_h, 1):
    ws7.cell(row=3, column=c, value=h)
style_header(ws7, 3, len(reg_h))

for r, reg in enumerate(regions, 4):
    ws7.cell(row=r, column=1, value=reg)
    ws7.cell(row=r, column=2, value=f"=SUMIF('Raw Data'!M2:M{LAST},A{r},'Raw Data'!Q2:Q{LAST})")
    ws7.cell(row=r, column=3, value=f"=SUMIF('Raw Data'!M2:M{LAST},A{r},'Raw Data'!T2:T{LAST})")
    ws7.cell(row=r, column=4, value=f'=IF(B{r}=0,0,C{r}/B{r})')
    ws7.cell(row=r, column=5, value=f"=COUNTIF('Raw Data'!M2:M{LAST},A{r})")
    ws7.cell(row=r, column=6, value=f'=IF(E{r}=0,0,B{r}/E{r})')
    ws7.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws7.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws7.cell(row=r, column=4).number_format = PCT_FMT
    ws7.cell(row=r, column=6).number_format = CURRENCY_FMT

tr = 8
ws7.cell(row=tr, column=1, value='TOTAL').font = BOLD_FONT
for c in [2, 3, 5]:
    cl = get_column_letter(c)
    ws7.cell(row=tr, column=c, value=f'=SUM({cl}4:{cl}7)')
    ws7.cell(row=tr, column=c).font = BOLD_FONT
ws7.cell(row=tr, column=4, value=f'=IF(B{tr}=0,0,C{tr}/B{tr})')
ws7.cell(row=tr, column=6, value=f'=IF(E{tr}=0,0,B{tr}/E{tr})')
ws7.cell(row=tr, column=2).number_format = CURRENCY_FMT
ws7.cell(row=tr, column=3).number_format = CURRENCY_FMT
ws7.cell(row=tr, column=4).number_format = PCT_FMT
ws7.cell(row=tr, column=6).number_format = CURRENCY_FMT
style_data_rows(ws7, 4, tr, len(reg_h))

# Bar chart BELOW table
chart_start_7 = tr + 3
bar_reg = BarChart()
bar_reg.type = "col"; bar_reg.title = "Sales & Profit by Region"; bar_reg.style = 10
bar_reg.width = 18; bar_reg.height = 12
bar_reg.add_data(Reference(ws7, min_col=2, min_row=3, max_row=7), titles_from_data=True)
bar_reg.add_data(Reference(ws7, min_col=3, min_row=3, max_row=7), titles_from_data=True)
bar_reg.set_categories(Reference(ws7, min_col=1, min_row=4, max_row=7))
ws7.add_chart(bar_reg, f"A{chart_start_7}")

# Top 15 States: after chart
state_start = chart_start_7 + CHART_ROWS_MED + 3
add_title(ws7, 'Top 15 States by Revenue', state_start, 1)
state_data = df.groupby('State').agg({'Sales':'sum','Profit':'sum','Order ID':'nunique'}).reset_index()
state_data = state_data.sort_values('Sales', ascending=False).head(15)

state_h = ['Rank', 'State', 'Total Sales', 'Total Profit', 'Profit Margin', 'Orders']
for c, h in enumerate(state_h, 1):
    ws7.cell(row=state_start + 2, column=c, value=h)
style_header(ws7, state_start + 2, len(state_h))

for r, (_, row) in enumerate(state_data.iterrows(), state_start + 3):
    ws7.cell(row=r, column=1, value=r - state_start - 2)
    ws7.cell(row=r, column=2, value=row['State'])
    ws7.cell(row=r, column=3, value=round(row['Sales'], 2))
    ws7.cell(row=r, column=4, value=round(row['Profit'], 2))
    ws7.cell(row=r, column=5, value=f'=IF(C{r}=0,0,D{r}/C{r})')
    ws7.cell(row=r, column=6, value=int(row['Order ID']))
    ws7.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws7.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws7.cell(row=r, column=5).number_format = PCT_FMT

state_end = state_start + 2 + 15
style_data_rows(ws7, state_start + 3, state_end, len(state_h))
set_col_widths(ws7, [10, 18, 16, 16, 14, 12])

# =====================================================================
# SHEET 8: TIME SERIES ANALYSIS
# =====================================================================
ws8 = wb.create_sheet('Time Series Analysis')
add_title(ws8, 'Time Series & Seasonality Analysis', 1, 1)

monthly_ts = df.groupby([df['Year'], df['Month']])['Sales'].sum().reset_index()
monthly_ts.columns = ['Year', 'Month', 'Sales']
monthly_ts = monthly_ts.sort_values(['Year', 'Month'])
monthly_ts['Period'] = monthly_ts['Year'].astype(int).astype(str) + '-' + monthly_ts['Month'].astype(int).astype(str).str.zfill(2)

ts_h = ['Period', 'Revenue', '3-Month MA', '6-Month MA', 'MoM Growth']
for c, h in enumerate(ts_h, 1):
    ws8.cell(row=3, column=c, value=h)
style_header(ws8, 3, len(ts_h))

for r, (_, row) in enumerate(monthly_ts.iterrows(), 4):
    ws8.cell(row=r, column=1, value=row['Period'])
    ws8.cell(row=r, column=2, value=round(row['Sales'], 2))
    ws8.cell(row=r, column=2).number_format = CURRENCY_FMT
    if r >= 6:
        ws8.cell(row=r, column=3, value=f'=AVERAGE(B{r-2}:B{r})')
        ws8.cell(row=r, column=3).number_format = CURRENCY_FMT
    if r >= 9:
        ws8.cell(row=r, column=4, value=f'=AVERAGE(B{r-5}:B{r})')
        ws8.cell(row=r, column=4).number_format = CURRENCY_FMT
    if r > 4:
        ws8.cell(row=r, column=5, value=f'=IF(B{r-1}=0,0,(B{r}-B{r-1})/B{r-1})')
        ws8.cell(row=r, column=5).number_format = PCT_FMT

ts_end = 3 + len(monthly_ts)
style_data_rows(ws8, 4, ts_end, len(ts_h))
set_col_widths(ws8, [12, 14, 14, 14, 12])

# Line chart BELOW table
chart_start_8 = ts_end + 3
line_ts = LineChart()
line_ts.title = "Revenue with Moving Averages"; line_ts.style = 10
line_ts.width = 26; line_ts.height = 14; line_ts.y_axis.title = "Revenue ($)"
for col_idx in [2, 3, 4]:
    line_ts.add_data(Reference(ws8, min_col=col_idx, min_row=3, max_row=ts_end), titles_from_data=True)
line_ts.set_categories(Reference(ws8, min_col=1, min_row=4, max_row=ts_end))
ws8.add_chart(line_ts, f"A{chart_start_8}")

# Seasonality: AFTER chart
season_start = chart_start_8 + CHART_ROWS_TALL + 3
add_title(ws8, 'Seasonality Analysis (Avg Monthly Revenue)', season_start, 1)
months_list = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
seas_h = ['Month', 'Month Name', 'Avg Revenue', 'Seasonality Index']
for c, h in enumerate(seas_h, 1):
    ws8.cell(row=season_start + 2, column=c, value=h)
style_header(ws8, season_start + 2, len(seas_h))

season = df.groupby('Month')['Sales'].mean()
seas_data_start = season_start + 3
seas_data_end = season_start + 14
for r, m in enumerate(range(1, 13), seas_data_start):
    ws8.cell(row=r, column=1, value=m)
    ws8.cell(row=r, column=2, value=months_list[m-1])
    ws8.cell(row=r, column=3, value=round(season.get(m, 0), 2))
    ws8.cell(row=r, column=3).number_format = CURRENCY_FMT
    avg_range = f'C{seas_data_start}:C{seas_data_end}'
    ws8.cell(row=r, column=4, value=f'=IF(AVERAGE({avg_range})=0,0,C{r}/AVERAGE({avg_range}))')
    ws8.cell(row=r, column=4).number_format = '0.000'

style_data_rows(ws8, seas_data_start, seas_data_end, len(seas_h))

# =====================================================================
# SHEET 9: SHIPPING ANALYSIS
# =====================================================================
ws9 = wb.create_sheet('Shipping Analysis')
add_title(ws9, 'Shipping & Delivery Analysis', 1, 1)

ship_modes = ['Standard Class', 'Second Class', 'First Class', 'Same Day']
ship_h = ['Ship Mode', 'Total Sales', 'Total Profit', 'Profit Margin', 'Order Count', '% of Orders', 'Avg Delivery Days']
for c, h in enumerate(ship_h, 1):
    ws9.cell(row=3, column=c, value=h)
style_header(ws9, 3, len(ship_h))

for r, mode in enumerate(ship_modes, 4):
    ws9.cell(row=r, column=1, value=mode)
    ws9.cell(row=r, column=2, value=f"=SUMIF('Raw Data'!E2:E{LAST},A{r},'Raw Data'!Q2:Q{LAST})")
    ws9.cell(row=r, column=3, value=f"=SUMIF('Raw Data'!E2:E{LAST},A{r},'Raw Data'!T2:T{LAST})")
    ws9.cell(row=r, column=4, value=f'=IF(B{r}=0,0,C{r}/B{r})')
    ws9.cell(row=r, column=5, value=f"=COUNTIF('Raw Data'!E2:E{LAST},A{r})")
    ws9.cell(row=r, column=6, value=f'=IF(SUM(E4:E7)=0,0,E{r}/SUM(E4:E7))')
    ws9.cell(row=r, column=7, value=f"=AVERAGEIF('Raw Data'!E2:E{LAST},A{r},'Raw Data'!U2:U{LAST})")
    ws9.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws9.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws9.cell(row=r, column=4).number_format = PCT_FMT
    ws9.cell(row=r, column=6).number_format = PCT_FMT
    ws9.cell(row=r, column=7).number_format = '0.0'

tr = 8
ws9.cell(row=tr, column=1, value='TOTAL').font = BOLD_FONT
for c in [2, 3, 5]:
    cl = get_column_letter(c)
    ws9.cell(row=tr, column=c, value=f'=SUM({cl}4:{cl}7)')
    ws9.cell(row=tr, column=c).font = BOLD_FONT
ws9.cell(row=tr, column=4, value=f'=IF(B{tr}=0,0,C{tr}/B{tr})')
ws9.cell(row=tr, column=7, value=f"=AVERAGE('Raw Data'!U2:U{LAST})")
ws9.cell(row=tr, column=2).number_format = CURRENCY_FMT
ws9.cell(row=tr, column=3).number_format = CURRENCY_FMT
ws9.cell(row=tr, column=4).number_format = PCT_FMT
ws9.cell(row=tr, column=7).number_format = '0.0'
style_data_rows(ws9, 4, tr, len(ship_h))

# Pie chart BELOW table
pie_start_9 = tr + 3
pie_ship = PieChart()
pie_ship.title = "Orders by Shipping Mode"; pie_ship.style = 10
pie_ship.width = 14; pie_ship.height = 11
pie_ship.add_data(Reference(ws9, min_col=5, min_row=4, max_row=7))
pie_ship.set_categories(Reference(ws9, min_col=1, min_row=4, max_row=7))
pie_ship.dataLabels = DataLabelList()
pie_ship.dataLabels.showPercent = True
ws9.add_chart(pie_ship, f"A{pie_start_9}")

# Delivery time table AFTER pie chart
del_start = pie_start_9 + CHART_ROWS_SHORT + 3
add_title(ws9, 'Delivery Time Distribution', del_start, 1)
delivery_bins = pd.cut(df['Delivery_Days'], bins=[0,2,4,6,8,100], labels=['0-2 days','3-4 days','5-6 days','7-8 days','8+ days'])
del_dist = delivery_bins.value_counts().sort_index()

del_h = ['Delivery Window', 'Order Count', '% of Total']
for c, h in enumerate(del_h, 1):
    ws9.cell(row=del_start + 2, column=c, value=h)
style_header(ws9, del_start + 2, len(del_h))

del_data_start = del_start + 3
del_data_end = del_start + 7
for r, (label, count) in enumerate(del_dist.items(), del_data_start):
    ws9.cell(row=r, column=1, value=str(label))
    ws9.cell(row=r, column=2, value=int(count))
    ws9.cell(row=r, column=3, value=f'=IF(SUM(B{del_data_start}:B{del_data_end})=0,0,B{r}/SUM(B{del_data_start}:B{del_data_end}))')
    ws9.cell(row=r, column=3).number_format = PCT_FMT

style_data_rows(ws9, del_data_start, del_data_end, len(del_h))
set_col_widths(ws9, [18, 14, 14, 14, 14, 12, 16])

# =====================================================================
# SHEET 10: DISCOUNT IMPACT
# =====================================================================
ws10 = wb.create_sheet('Discount Impact')
add_title(ws10, 'Discount Impact on Profitability', 1, 1)

df['Discount_Band'] = pd.cut(df['Discount'], bins=[-0.01, 0, 0.1, 0.2, 0.3, 0.5, 1.0],
                              labels=['No Discount', '1-10%', '11-20%', '21-30%', '31-50%', '50%+'])
disc_analysis = df.groupby('Discount_Band', observed=True).agg({
    'Sales': ['sum', 'mean', 'count'], 'Profit': ['sum', 'mean'],
}).reset_index()
disc_analysis.columns = ['Discount Band', 'Total Sales', 'Avg Sale', 'Order Count', 'Total Profit', 'Avg Profit']

disc_h = ['Discount Band', 'Order Count', '% of Orders', 'Total Sales', 'Total Profit', 'Avg Profit/Order', 'Profit Margin']
for c, h in enumerate(disc_h, 1):
    ws10.cell(row=3, column=c, value=h)
style_header(ws10, 3, len(disc_h))

disc_data_start = 4
disc_data_end = 3 + len(disc_analysis)
for r, (_, row) in enumerate(disc_analysis.iterrows(), disc_data_start):
    ws10.cell(row=r, column=1, value=str(row['Discount Band']))
    ws10.cell(row=r, column=2, value=int(row['Order Count']))
    ws10.cell(row=r, column=4, value=round(row['Total Sales'], 2))
    ws10.cell(row=r, column=5, value=round(row['Total Profit'], 2))
    ws10.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws10.cell(row=r, column=5).number_format = CURRENCY_FMT
    if row['Total Profit'] < 0:
        ws10.cell(row=r, column=5).font = RED_FONT

for r in range(disc_data_start, disc_data_end + 1):
    total_o = f'SUM(B{disc_data_start}:B{disc_data_end})'
    ws10.cell(row=r, column=3, value=f'=IF({total_o}=0,0,B{r}/{total_o})')
    ws10.cell(row=r, column=3).number_format = PCT_FMT
    ws10.cell(row=r, column=6, value=f'=IF(B{r}=0,0,E{r}/B{r})')
    ws10.cell(row=r, column=6).number_format = CURRENCY_FMT
    ws10.cell(row=r, column=7, value=f'=IF(D{r}=0,0,E{r}/D{r})')
    ws10.cell(row=r, column=7).number_format = PCT_FMT

style_data_rows(ws10, disc_data_start, disc_data_end, len(disc_h))
set_col_widths(ws10, [16, 14, 12, 14, 14, 16, 14])

# Bar chart BELOW table
chart_start_10 = disc_data_end + 3
bar_disc = BarChart()
bar_disc.type = "col"; bar_disc.title = "Profit Margin by Discount Band"; bar_disc.style = 10
bar_disc.width = 20; bar_disc.height = 12
bar_disc.add_data(Reference(ws10, min_col=7, min_row=3, max_row=disc_data_end), titles_from_data=True)
bar_disc.set_categories(Reference(ws10, min_col=1, min_row=disc_data_start, max_row=disc_data_end))
ws10.add_chart(bar_disc, f"A{chart_start_10}")

# Category x Discount: AFTER chart
cat_disc_start = chart_start_10 + CHART_ROWS_MED + 3
add_title(ws10, 'Discount Impact by Category', cat_disc_start, 1)
cat_disc = df.groupby(['Category', 'Discount_Band'], observed=True).agg({'Sales':'sum','Profit':'sum'}).reset_index()

catd_h = ['Category', 'Discount Band', 'Total Sales', 'Total Profit', 'Profit Margin']
for c, h in enumerate(catd_h, 1):
    ws10.cell(row=cat_disc_start + 2, column=c, value=h)
style_header(ws10, cat_disc_start + 2, len(catd_h))

for r, (_, row) in enumerate(cat_disc.iterrows(), cat_disc_start + 3):
    ws10.cell(row=r, column=1, value=row['Category'])
    ws10.cell(row=r, column=2, value=str(row['Discount_Band']))
    ws10.cell(row=r, column=3, value=round(row['Sales'], 2))
    ws10.cell(row=r, column=4, value=round(row['Profit'], 2))
    ws10.cell(row=r, column=5, value=f'=IF(C{r}=0,0,D{r}/C{r})')
    ws10.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws10.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws10.cell(row=r, column=5).number_format = PCT_FMT

catd_end = cat_disc_start + 2 + len(cat_disc)
style_data_rows(ws10, cat_disc_start + 3, catd_end, len(catd_h))

# =====================================================================
# SHEET 11: KPI DASHBOARD
# =====================================================================
ws11 = wb.create_sheet('KPI Dashboard')
add_title(ws11, 'Executive KPI Dashboard', 1, 1)
ws11.cell(row=2, column=1, value=f'Data Period: {df["Order Date"].min().strftime("%b %Y")} - {df["Order Date"].max().strftime("%b %Y")}').font = Font(name='Arial', size=10, italic=True, color='666666')

kpi_labels = [
    ('Total Revenue', f"=SUM('Raw Data'!Q2:Q{LAST})", CURRENCY_FMT),
    ('Total Profit', f"=SUM('Raw Data'!T2:T{LAST})", CURRENCY_FMT),
    ('Profit Margin', f"=SUM('Raw Data'!T2:T{LAST})/SUM('Raw Data'!Q2:Q{LAST})", PCT_FMT),
    ('Total Orders', f"=COUNTA('Raw Data'!B2:B{LAST})", NUM_FMT),
    ('Total Units Sold', f"=SUM('Raw Data'!R2:R{LAST})", NUM_FMT),
    ('Avg Order Value', f"=SUM('Raw Data'!Q2:Q{LAST})/COUNTA('Raw Data'!B2:B{LAST})", CURRENCY_FMT),
    ('Avg Profit/Order', f"=SUM('Raw Data'!T2:T{LAST})/COUNTA('Raw Data'!B2:B{LAST})", CURRENCY_FMT),
    ('Avg Discount', f"=AVERAGE('Raw Data'!S2:S{LAST})", PCT_FMT),
    ('Avg Delivery Days', f"=AVERAGE('Raw Data'!U2:U{LAST})", '0.0'),
]

for i, (label, formula, fmt) in enumerate(kpi_labels):
    row_offset = 4 + (i // 3) * 3
    col_offset = 1 + (i % 3) * 3
    cell_label = ws11.cell(row=row_offset, column=col_offset, value=label)
    cell_label.font = Font(name='Arial', size=9, color='FFFFFF', bold=True)
    cell_label.fill = KPI_FILL
    cell_label.alignment = Alignment(horizontal='center')
    ws11.merge_cells(start_row=row_offset, start_column=col_offset, end_row=row_offset, end_column=col_offset+1)
    cell_val = ws11.cell(row=row_offset + 1, column=col_offset, value=formula)
    cell_val.font = Font(name='Arial', size=16, bold=True, color='1F4E79')
    cell_val.alignment = Alignment(horizontal='center')
    cell_val.number_format = fmt
    ws11.merge_cells(start_row=row_offset+1, start_column=col_offset, end_row=row_offset+1, end_column=col_offset+1)

# Category summary
cat_kpi_row = 16
add_title(ws11, 'Category Performance', cat_kpi_row, 1)
cat_kpi_h = ['Category', 'Revenue', 'Profit', 'Margin', 'Orders', '% of Revenue']
for c, h in enumerate(cat_kpi_h, 1):
    ws11.cell(row=cat_kpi_row + 2, column=c, value=h)
style_header(ws11, cat_kpi_row + 2, len(cat_kpi_h))

for r, cat in enumerate(categories, cat_kpi_row + 3):
    ws11.cell(row=r, column=1, value=cat)
    ws11.cell(row=r, column=2, value=f"=SUMIF('Raw Data'!N2:N{LAST},A{r},'Raw Data'!Q2:Q{LAST})")
    ws11.cell(row=r, column=3, value=f"=SUMIF('Raw Data'!N2:N{LAST},A{r},'Raw Data'!T2:T{LAST})")
    ws11.cell(row=r, column=4, value=f'=IF(B{r}=0,0,C{r}/B{r})')
    ws11.cell(row=r, column=5, value=f"=COUNTIF('Raw Data'!N2:N{LAST},A{r})")
    ws11.cell(row=r, column=6, value=f"=IF(SUM('Raw Data'!Q2:Q{LAST})=0,0,B{r}/SUM('Raw Data'!Q2:Q{LAST}))")
    ws11.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws11.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws11.cell(row=r, column=4).number_format = PCT_FMT
    ws11.cell(row=r, column=6).number_format = PCT_FMT
style_data_rows(ws11, cat_kpi_row + 3, cat_kpi_row + 5, len(cat_kpi_h))

# Regional summary
reg_kpi_row = cat_kpi_row + 8
add_title(ws11, 'Regional Performance', reg_kpi_row, 1)
reg_kpi_h = ['Region', 'Revenue', 'Profit', 'Margin', 'Orders', '% of Revenue']
for c, h in enumerate(reg_kpi_h, 1):
    ws11.cell(row=reg_kpi_row + 2, column=c, value=h)
style_header(ws11, reg_kpi_row + 2, 6)

for r, reg in enumerate(regions, reg_kpi_row + 3):
    ws11.cell(row=r, column=1, value=reg)
    ws11.cell(row=r, column=2, value=f"=SUMIF('Raw Data'!M2:M{LAST},A{r},'Raw Data'!Q2:Q{LAST})")
    ws11.cell(row=r, column=3, value=f"=SUMIF('Raw Data'!M2:M{LAST},A{r},'Raw Data'!T2:T{LAST})")
    ws11.cell(row=r, column=4, value=f'=IF(B{r}=0,0,C{r}/B{r})')
    ws11.cell(row=r, column=5, value=f"=COUNTIF('Raw Data'!M2:M{LAST},A{r})")
    ws11.cell(row=r, column=6, value=f"=IF(SUM('Raw Data'!Q2:Q{LAST})=0,0,B{r}/SUM('Raw Data'!Q2:Q{LAST}))")
    ws11.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws11.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws11.cell(row=r, column=4).number_format = PCT_FMT
    ws11.cell(row=r, column=6).number_format = PCT_FMT
style_data_rows(ws11, reg_kpi_row + 3, reg_kpi_row + 6, 6)

for col in range(1, 10):
    ws11.column_dimensions[get_column_letter(col)].width = 16

# =====================================================================
# SHEET 12: COHORT ANALYSIS
# =====================================================================
ws12 = wb.create_sheet('Cohort Analysis')
add_title(ws12, 'Customer Cohort Retention Analysis', 1, 1)

df['Order_Month'] = df['Order Date'].dt.to_period('M')
first_purchase = df.groupby('Customer ID')['Order_Month'].min().reset_index()
first_purchase.columns = ['Customer ID', 'Cohort']
df_cohort = df.merge(first_purchase, on='Customer ID')
df_cohort['Cohort_Index'] = (df_cohort['Order_Month'] - df_cohort['Cohort']).apply(lambda x: x.n)

cohort_data = df_cohort.groupby(['Cohort', 'Cohort_Index'])['Customer ID'].nunique().reset_index()
cohort_pivot = cohort_data.pivot(index='Cohort', columns='Cohort_Index', values='Customer ID').fillna(0)
cohort_sizes = cohort_pivot[0]
cohort_pct = cohort_pivot.div(cohort_sizes, axis=0) * 100

quarterly_cohorts = {}
for cohort in cohort_pct.index:
    q = f"{cohort.year} Q{(cohort.month-1)//3+1}"
    quarterly_cohorts.setdefault(q, []).append(cohort)

add_title(ws12, 'Quarterly Cohort Retention (%)', 3, 1)
q_periods = ['Month 0', 'Month 3', 'Month 6', 'Month 9', 'Month 12', 'Month 18', 'Month 24']
coh_h = ['Cohort Quarter', 'New Customers'] + q_periods
for c, h in enumerate(coh_h, 1):
    ws12.cell(row=5, column=c, value=h)
style_header(ws12, 5, len(coh_h))

row_idx = 6
for q_name, cohorts in sorted(quarterly_cohorts.items()):
    total_new = sum(cohort_sizes.get(c, 0) for c in cohorts)
    ws12.cell(row=row_idx, column=1, value=q_name)
    ws12.cell(row=row_idx, column=2, value=int(total_new))
    for ci, month_idx in enumerate([0, 3, 6, 9, 12, 18, 24]):
        vals = []
        for c in cohorts:
            if month_idx in cohort_pct.columns and c in cohort_pct.index:
                v = cohort_pct.loc[c, month_idx]
                if v > 0:
                    vals.append(v)
        if vals:
            ws12.cell(row=row_idx, column=3+ci, value=round(np.mean(vals), 1))
            ws12.cell(row=row_idx, column=3+ci).number_format = '0.0"%"'
    row_idx += 1

style_data_rows(ws12, 6, row_idx - 1, len(coh_h))

# Revenue by cohort
rev_start = row_idx + 2
add_title(ws12, 'Revenue by Customer Cohort Quarter', rev_start, 1)
rev_cohort = df_cohort.groupby(['Cohort', 'Cohort_Index'])['Sales'].sum().reset_index()
rev_pivot = rev_cohort.pivot(index='Cohort', columns='Cohort_Index', values='Sales').fillna(0)

rev_h = ['Cohort Quarter', 'Total Revenue', 'Avg Revenue/Customer', 'Customer Count']
for c, h in enumerate(rev_h, 1):
    ws12.cell(row=rev_start + 2, column=c, value=h)
style_header(ws12, rev_start + 2, len(rev_h))

rev_row = rev_start + 3
for q_name, cohorts in sorted(quarterly_cohorts.items()):
    total_rev_q = sum(rev_pivot.loc[c].sum() if c in rev_pivot.index else 0 for c in cohorts)
    total_cust = sum(cohort_sizes.get(c, 0) for c in cohorts)
    ws12.cell(row=rev_row, column=1, value=q_name)
    ws12.cell(row=rev_row, column=2, value=round(total_rev_q, 2))
    ws12.cell(row=rev_row, column=2).number_format = CURRENCY_FMT
    ws12.cell(row=rev_row, column=3, value=f'=IF(D{rev_row}=0,0,B{rev_row}/D{rev_row})')
    ws12.cell(row=rev_row, column=3).number_format = CURRENCY_FMT
    ws12.cell(row=rev_row, column=4, value=int(total_cust))
    rev_row += 1

style_data_rows(ws12, rev_start + 3, rev_row - 1, len(rev_h))
set_col_widths(ws12, [16, 14] + [12]*7)

# =====================================================================
# SAVE
# =====================================================================
output_path = 'output/Sales_Analysis_Report.xlsx'
wb.save(output_path)
print(f'Workbook saved: {output_path}')
print(f'Sheets ({len(wb.sheetnames)}): {", ".join(wb.sheetnames)}')
print(f'Raw Data: {N} rows | RFM customers: {n_cust}')
